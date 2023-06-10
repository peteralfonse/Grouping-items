#!/usr/bin/env python
# coding: utf-8

# In[13]:


import openpyxl


# 'import openpyxl' is a Python statement that allows you to access the functionalities of the openpyxl library for working with Excel files.

# In[14]:


workbook = openpyxl.load_workbook('file path')
sheet = workbook.active


# loading an Excel file and assigns it to the workbook variable, and sheet = workbook.active selects the active sheet within the workbook and assigns it to the sheet variable for further manipulation.

# In[15]:


my_dict = {key: [] for key in range(1, 1000000)}


# creating a dictionary called my_dict with keys ranging from 1 to 999,999, where each key is initially associated with an empty list as its value.

# In[16]:


for i in range(0,100):
        # Iterate through the values in the first column (excluding the first row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        value = row[0]
        if value is not None:
            value = str(value)  # Convert the value to a string for comparison

            # Find the next empty key in the dictionary
            for key, lst in my_dict.items():
                if not lst:
                    lst.append(value)
                    if row[1] is not None:
                        lst.append(row[1])  # Second column value
                    if row[2] is not None:
                        lst.append(row[2])  # Third column value
                    break
                elif value in lst:
                    if row[1] is not None and row[1] not in lst:
                        lst.append(row[1])  # Second column value
                    if row[2] is not None and row[2] not in lst:
                        lst.append(row[2])  # Third column value
                    break


# In this step i wanted to allocate each item and it's variables on a single key in the created dictionary

# In[17]:


filtered_dict = {}


# In[18]:


# Iterate through the keys and lists in the original dictionary
for key, lst in my_dict.items():
    # Check if the current list is a subset of any other list in the dictionary
    is_subset = any(set(lst) <= set(other_lst) for other_lst in my_dict.values() if other_lst != lst)
    
    if not is_subset:
        # Add the current key and list to the filtered dictionary
        filtered_dict[key] = lst


# In[19]:


print(filtered_dict)


# Here i wanted to clean my dictionary from any excessive information to avoid confusion

# In[20]:


# Add a new column header
sheet.cell(row=1, column=4).value = 'Reference'


# In[21]:


for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    value = row[0].value
    if value is not None:
        value = str(value)
        for key, lst in my_dict.items():
            if value in lst:
                # Add the key value to the fourth column
                sheet.cell(row=row[0].row, column=4).value = key
                break


# Here i wanted to assign each item to it's coressponding index

# In[22]:


workbook.save('Reference')


# In[23]:


file_path = 'file path'


# In[24]:


workbook.save(file_path)


# In[25]:


print(f"The modified Excel file has been saved as: {file_path}")


# In[ ]:




